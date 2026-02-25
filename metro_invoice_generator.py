#!/usr/bin/env python3
"""
=============================================================================
GÉNÉRATEUR DE FACTURES METRO → EXCEL
=============================================================================
Version: 3.0 — Production Ready
Fournisseurs supportés: METRO (avec alcools, cotis sécu, remises APPM)
Aussi compatible: Carrefour, Codifrance (structure adaptable)

LOGIQUE METRO:
- Les articles sont groupés par SECTION (BRASSERIE, SPIRITUEUX, CAVE, etc.)
- Chaque section a un sous-total affiché sur la facture ("*** SECTION Total: X.XX")
- Les COTISATIONS SECURITE SOCIALE sont des lignes séparées ajoutées au total section
- Les REMISES "Achetez Plus Payez Moins" sont des lignes négatives
- Le Total HT = somme sections - remises globales
- 3 taux TVA possibles: B=5.50%, C=20.00%, D=20.00%
- Total TTC = Total HT + Total TVA

STRUCTURE D'UNE LIGNE ARTICLE METRO:
  EAN | N°Article | Désignation | Régie | Vol% | VAP | Poids/VE | Prix Unit. | Colis | Qté | Montant | TVA | Promo
  
PARTICULARITÉS:
- Articles alcool (TVA D): ont souvent "Plus : COTIS. SECURITE SOCIALE X.XX D"
- Articles avec lot: "N° GTIN et LOT : ... LOT-Nr:..."
- Articles avec DLC: "_BEST_BEFORE_DATEYYMMDD"
- Remises volume: "Offre Achetez Plus Payez Moins X.XX-"
- Promo "3 POUR 2": remise = prix unitaire négatif
- Le sous-total section INCLUT les cotis mais EXCLUT les remises APPM
- Le Total HT EXCLUT les remises APPM (elles sont déduites globalement)

VÉRIFICATION OBLIGATOIRE AVANT GÉNÉRATION:
1. Somme des montants articles par section = sous-total section (± cotis)
2. Somme sections - remises = Total HT facture
3. Total HT × taux TVA = Montant TVA par taux
4. Total HT + Total TVA = Total TTC
→ Si UN SEUL de ces checks échoue → NE PAS GÉNÉRER, afficher l'erreur

=============================================================================
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.properties import PageSetupProperties
from copy import copy
import sys
import json
from decimal import Decimal, ROUND_HALF_UP

# =============================================================================
# CONFIGURATION COULEURS & STYLES
# =============================================================================
COLORS = {
    "DARK_BLUE": "1F3864",
    "MEDIUM_BLUE": "2E75B6",
    "LIGHT_BLUE": "D6E4F0",
    "WHITE": "FFFFFF",
    "LIGHT_GRAY": "F2F2F2",
    "ORANGE": "ED7D31",
    "GREEN_DARK": "548235",
    "RED": "C00000",
    "PURPLE": "7030A0",
    "DARK_GRAY": "666666",
}

THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# =============================================================================
# CLASSES DE DONNÉES
# =============================================================================

class InvoiceItem:
    """Un article de la facture."""
    def __init__(self, ean, ref, designation, prix_unit, colisage, qte, montant, tva_code, cotis_secu=0, promo=""):
        self.ean = ean
        self.ref = ref
        self.designation = designation
        self.prix_unit = prix_unit
        self.colisage = colisage
        self.qte = qte
        self.montant = montant
        self.tva_code = tva_code
        self.cotis_secu = cotis_secu
        self.promo = promo

class DiscountLine:
    """Une ligne de remise."""
    def __init__(self, description, montant, tva_code="D"):
        self.description = description
        self.montant = montant  # Négatif
        self.tva_code = tva_code

class Section:
    """Une section de la facture (ex: BRASSERIE, SPIRITUEUX, etc.)"""
    def __init__(self, name, expected_total=None):
        self.name = name
        self.items = []          # Liste de InvoiceItem
        self.discounts = []      # Liste de DiscountLine
        self.expected_total = expected_total  # Total affiché sur la facture

    def computed_total_before_discounts(self):
        """Somme articles + cotis (comme affiché dans les sous-totaux Metro)."""
        return sum(i.montant for i in self.items) + sum(i.cotis_secu for i in self.items)

    def computed_total_with_discounts(self):
        """Total net incluant remises."""
        return self.computed_total_before_discounts() + sum(d.montant for d in self.discounts)

class TVARate:
    """Un taux de TVA."""
    def __init__(self, code, rate, base_ht, montant_tva, montant_ttc):
        self.code = code
        self.rate = rate
        self.base_ht = base_ht
        self.montant_tva = montant_tva
        self.montant_ttc = montant_ttc

class MetroInvoice:
    """Facture METRO complète."""
    def __init__(self):
        self.numero = ""
        self.date = ""
        self.client_name = ""
        self.client_address = ""
        self.client_num = ""
        self.magasin = ""
        self.sections = []
        self.tva_rates = []
        self.total_ht = 0
        self.total_tva = 0
        self.total_ttc = 0
        self.total_cotis = 0
        self.payment_info = ""
        self.nb_colis = 0
        self.notes = ""  # Ex: "Mise en attente rappelée et facturée..."

    def all_items(self):
        """Tous les articles de toutes les sections."""
        items = []
        for s in self.sections:
            items.extend(s.items)
        return items

    def all_discounts(self):
        """Toutes les remises de toutes les sections."""
        discounts = []
        for s in self.sections:
            discounts.extend(s.discounts)
        return discounts

    def total_discounts(self):
        """Somme de toutes les remises."""
        return sum(d.montant for d in self.all_discounts())

    def sum_section_totals(self):
        """Somme des totaux sections (avant remises globales)."""
        return sum(s.computed_total_before_discounts() for s in self.sections)

    def computed_total_ht(self):
        """Total HT calculé = sections - remises."""
        return self.sum_section_totals() + self.total_discounts()


# =============================================================================
# VÉRIFICATION DES TOTAUX
# =============================================================================

def verify_invoice(invoice):
    """
    Vérifie que tous les totaux de la facture sont cohérents.
    Retourne (success, errors_list).
    Si success=False, NE PAS GÉNÉRER le fichier.
    """
    errors = []
    warnings = []

    # 1. Vérifier chaque sous-total section
    for section in invoice.sections:
        if section.expected_total is not None:
            computed = section.computed_total_before_discounts()
            diff = abs(computed - section.expected_total)
            if diff > 0.02:
                errors.append(
                    f"SECTION '{section.name}': calculé={computed:.2f}, "
                    f"attendu={section.expected_total:.2f}, diff={diff:.2f}"
                )
            else:
                print(f"  ✓ Section {section.name}: {computed:.2f} = {section.expected_total:.2f}")

    # 2. Vérifier Total HT
    computed_ht = invoice.computed_total_ht()
    diff_ht = abs(computed_ht - invoice.total_ht)
    if diff_ht > 0.02:
        errors.append(
            f"TOTAL HT: calculé={computed_ht:.2f}, attendu={invoice.total_ht:.2f}, diff={diff_ht:.2f}"
        )
    else:
        print(f"  ✓ Total HT: {computed_ht:.2f} = {invoice.total_ht:.2f}")

    # 3. Vérifier Total TVA
    computed_tva = sum(t.montant_tva for t in invoice.tva_rates)
    diff_tva = abs(computed_tva - invoice.total_tva)
    if diff_tva > 0.02:
        errors.append(
            f"TOTAL TVA: calculé={computed_tva:.2f}, attendu={invoice.total_tva:.2f}, diff={diff_tva:.2f}"
        )
    else:
        print(f"  ✓ Total TVA: {computed_tva:.2f} = {invoice.total_tva:.2f}")

    # 4. Vérifier Total TTC
    computed_ttc = invoice.total_ht + invoice.total_tva
    diff_ttc = abs(computed_ttc - invoice.total_ttc)
    if diff_ttc > 0.02:
        errors.append(
            f"TOTAL TTC: calculé={computed_ttc:.2f}, attendu={invoice.total_ttc:.2f}, diff={diff_ttc:.2f}"
        )
    else:
        print(f"  ✓ Total TTC: {computed_ttc:.2f} = {invoice.total_ttc:.2f}")

    # 5. Vérifier bases TVA
    sum_bases = sum(t.base_ht for t in invoice.tva_rates)
    diff_bases = abs(sum_bases - invoice.total_ht)
    if diff_bases > 0.02:
        errors.append(
            f"BASES TVA: somme={sum_bases:.2f}, total_ht={invoice.total_ht:.2f}, diff={diff_bases:.2f}"
        )
    else:
        print(f"  ✓ Bases TVA: {sum_bases:.2f} = {invoice.total_ht:.2f}")

    # 6. Vérifier cotis sécu si applicable
    if invoice.total_cotis > 0:
        computed_cotis = sum(i.cotis_secu for i in invoice.all_items())
        diff_cotis = abs(computed_cotis - invoice.total_cotis)
        if diff_cotis > 0.02:
            errors.append(
                f"COTIS SECU: calculé={computed_cotis:.2f}, attendu={invoice.total_cotis:.2f}"
            )
        else:
            print(f"  ✓ Cotis Sécu: {computed_cotis:.2f} = {invoice.total_cotis:.2f}")

    if errors:
        print("\n❌ ERREURS DE VÉRIFICATION:")
        for e in errors:
            print(f"  ✗ {e}")
        return False, errors

    print("\n✅ TOUTES LES VÉRIFICATIONS PASSENT")
    return True, []


# =============================================================================
# GÉNÉRATEUR EXCEL
# =============================================================================

def generate_excel(invoice, output_path):
    """Génère le fichier Excel à partir d'une facture vérifiée."""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facture METRO"

    # Styles
    header_font = Font(name="Arial", bold=True, color=COLORS["WHITE"], size=10)
    header_fill = PatternFill("solid", fgColor=COLORS["DARK_BLUE"])
    section_font = Font(name="Arial", bold=True, color=COLORS["WHITE"], size=10)
    section_fill = PatternFill("solid", fgColor=COLORS["MEDIUM_BLUE"])
    data_font = Font(name="Arial", size=8.5)
    subtotal_font = Font(name="Arial", bold=True, size=9)
    subtotal_fill = PatternFill("solid", fgColor=COLORS["LIGHT_BLUE"])
    alt_fill = PatternFill("solid", fgColor=COLORS["LIGHT_GRAY"])
    cotis_font = Font(name="Arial", size=8, color=COLORS["PURPLE"], italic=True)
    remise_font = Font(name="Arial", size=8.5, color=COLORS["RED"], bold=True)
    total_font = Font(name="Arial", bold=True, color=COLORS["WHITE"], size=12)
    total_fill = PatternFill("solid", fgColor=COLORS["DARK_BLUE"])
    promo_font = Font(name="Arial", size=8, color=COLORS["ORANGE"], bold=True)
    green_fill = PatternFill("solid", fgColor=COLORS["GREEN_DARK"])
    green_font = Font(name="Arial", bold=True, color=COLORS["WHITE"], size=14)

    # Déterminer si on a des cotis (facture alcool)
    has_cotis = invoice.total_cotis > 0
    if has_cotis:
        col_widths = {'A': 7, 'B': 16, 'C': 11, 'D': 42, 'E': 9, 'F': 7, 'G': 6, 'H': 11, 'I': 6, 'J': 11, 'K': 6}
        headers = ["N°", "EAN", "Réf", "Désignation", "Prix Unit.", "Colis.", "Qté", "Montant HT", "TVA", "Cotis. Sécu", "Promo"]
        max_col = 11
    else:
        col_widths = {'A': 7, 'B': 16, 'C': 11, 'D': 45, 'E': 10, 'F': 8, 'G': 6, 'H': 12, 'I': 6, 'J': 6}
        headers = ["N°", "EAN", "Réf", "Désignation", "Prix Unit.", "Colis.", "Qté", "Montant HT", "TVA", "Promo"]
        max_col = 10

    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # Titre
    row = 1
    last_col_letter = chr(64 + max_col)
    ws.merge_cells(f'A{row}:{last_col_letter}{row}')
    ws['A1'] = f"FACTURE {invoice.magasin}"
    ws['A1'].font = Font(name="Arial", bold=True, size=14, color=COLORS["DARK_BLUE"])
    ws['A1'].alignment = Alignment(horizontal='center')

    row = 2
    ws.merge_cells(f'A{row}:{last_col_letter}{row}')
    subtitle = f"N° {invoice.numero} — Date: {invoice.date} — Client: {invoice.client_name}"
    if invoice.notes:
        subtitle += f" — {invoice.notes}"
    ws[f'A{row}'] = subtitle
    ws[f'A{row}'].font = Font(name="Arial", size=9, color=COLORS["DARK_GRAY"])
    ws[f'A{row}'].alignment = Alignment(horizontal='center')

    # Headers
    row = 4
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

    row += 1
    item_num = 0

    for section in invoice.sections:
        # Section header
        ws.merge_cells(f'A{row}:{last_col_letter}{row}')
        cell = ws.cell(row=row, column=1, value=f"  {section.name}")
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal='left')
        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).fill = section_fill
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

        # Items
        for item in section.items:
            item_num += 1
            is_alt = item_num % 2 == 0
            row_fill = alt_fill if is_alt else None

            if has_cotis:
                vals = [item_num, item.ean, item.ref, item.designation, item.prix_unit,
                        item.colisage, item.qte, item.montant, item.tva_code,
                        item.cotis_secu if item.cotis_secu > 0 else "", item.promo or ""]
            else:
                vals = [item_num, item.ean, item.ref, item.designation, item.prix_unit,
                        item.colisage, item.qte, item.montant, item.tva_code, item.promo or ""]

            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font = data_font
                cell.border = THIN_BORDER
                if row_fill:
                    cell.fill = row_fill
                if c == 1:
                    cell.alignment = Alignment(horizontal='center')
                elif c in (5, 8):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                elif c == 10 and has_cotis:
                    if v and v != "":
                        cell.number_format = '#,##0.00'
                        cell.font = cotis_font
                    cell.alignment = Alignment(horizontal='right')
                elif c in (6, 7):
                    cell.alignment = Alignment(horizontal='center')
                elif c == 9:
                    cell.alignment = Alignment(horizontal='center')
                elif c == max_col:
                    if v == "P":
                        cell.font = promo_font
                    cell.alignment = Alignment(horizontal='center')
            row += 1

        # Discount lines
        for discount in section.discounts:
            ws.cell(row=row, column=4, value=discount.description).font = remise_font
            cell_h = ws.cell(row=row, column=8, value=discount.montant)
            cell_h.font = remise_font
            cell_h.number_format = '#,##0.00'
            cell_h.alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=9, value=discount.tva_code).font = Font(name="Arial", size=8, color=COLORS["RED"])
            ws.cell(row=row, column=9).alignment = Alignment(horizontal='center')
            for c in range(1, max_col + 1):
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1

    # ========================================
    # RÉCAPITULATIF
    # ========================================
    row += 1
    ws.merge_cells(f'A{row}:{last_col_letter}{row}')
    ws.cell(row=row, column=1, value="RÉCAPITULATIF PAR SECTION").font = Font(
        name="Arial", bold=True, size=11, color=COLORS["DARK_BLUE"])
    row += 1

    for section in invoice.sections:
        total = section.expected_total if section.expected_total else section.computed_total_before_discounts()
        ws.merge_cells(f'A{row}:G{row}')
        cell = ws.cell(row=row, column=1, value=f"  {section.name}")
        cell.font = subtotal_font
        cell.fill = subtotal_fill
        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).fill = subtotal_fill
            ws.cell(row=row, column=c).border = THIN_BORDER
        cell_val = ws.cell(row=row, column=8, value=total)
        cell_val.font = subtotal_font
        cell_val.fill = subtotal_fill
        cell_val.number_format = '#,##0.00'
        cell_val.alignment = Alignment(horizontal='right')
        row += 1

    # Remises globales
    total_disc = invoice.total_discounts()
    if total_disc != 0:
        ws.merge_cells(f'A{row}:G{row}')
        cell = ws.cell(row=row, column=1, value="  REMISES")
        cell.font = Font(name="Arial", bold=True, size=10, color=COLORS["RED"])
        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).border = THIN_BORDER
        ws.cell(row=row, column=8, value=total_disc).font = Font(
            name="Arial", bold=True, size=10, color=COLORS["RED"])
        ws.cell(row=row, column=8).number_format = '#,##0.00'
        ws.cell(row=row, column=8).alignment = Alignment(horizontal='right')
        row += 1

    # Cotis info
    if has_cotis:
        ws.merge_cells(f'A{row}:G{row}')
        ws.cell(row=row, column=1, value="  Dont COTIS. SECURITE SOCIALE (incluse)").font = Font(
            name="Arial", size=9, color=COLORS["PURPLE"], italic=True)
        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).border = THIN_BORDER
        ws.cell(row=row, column=10, value=invoice.total_cotis).font = Font(
            name="Arial", size=9, color=COLORS["PURPLE"], bold=True)
        ws.cell(row=row, column=10).number_format = '#,##0.00'
        ws.cell(row=row, column=10).alignment = Alignment(horizontal='right')
        row += 1

    row += 1

    # ========================================
    # TOTAL HT
    # ========================================
    ws.merge_cells(f'A{row}:G{row}')
    cell = ws.cell(row=row, column=1, value="TOTAL HT")
    cell.font = total_font
    cell.fill = total_fill
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).fill = total_fill
        ws.cell(row=row, column=c).border = THIN_BORDER
    cell_val = ws.cell(row=row, column=8, value=invoice.total_ht)
    cell_val.font = total_font
    cell_val.fill = total_fill
    cell_val.number_format = '#,##0.00'
    cell_val.alignment = Alignment(horizontal='right')
    row += 1

    # TVA détails
    for tva in invoice.tva_rates:
        ws.merge_cells(f'A{row}:G{row}')
        ws.cell(row=row, column=1,
                value=f"  TVA {tva.code} ({tva.rate}%) — Base: {tva.base_ht:,.2f} €").font = Font(name="Arial", size=9)
        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).border = THIN_BORDER
        ws.cell(row=row, column=8, value=tva.montant_tva).font = Font(name="Arial", size=9)
        ws.cell(row=row, column=8).number_format = '#,##0.00'
        ws.cell(row=row, column=8).alignment = Alignment(horizontal='right')
        row += 1

    # Total TVA
    ws.merge_cells(f'A{row}:G{row}')
    ws.cell(row=row, column=1, value="TOTAL TVA").font = Font(
        name="Arial", bold=True, size=11, color=COLORS["MEDIUM_BLUE"])
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).border = THIN_BORDER
    ws.cell(row=row, column=8, value=invoice.total_tva).font = Font(
        name="Arial", bold=True, size=11, color=COLORS["MEDIUM_BLUE"])
    ws.cell(row=row, column=8).number_format = '#,##0.00'
    ws.cell(row=row, column=8).alignment = Alignment(horizontal='right')
    row += 2

    # ========================================
    # TOTAL TTC
    # ========================================
    ws.merge_cells(f'A{row}:G{row}')
    cell = ws.cell(row=row, column=1, value="TOTAL TTC À PAYER")
    cell.font = green_font
    cell.fill = green_fill
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).fill = green_fill
        ws.cell(row=row, column=c).border = THIN_BORDER
    cell_val = ws.cell(row=row, column=8, value=invoice.total_ttc)
    cell_val.font = green_font
    cell_val.fill = green_fill
    cell_val.number_format = '#,##0.00'
    cell_val.alignment = Alignment(horizontal='right')
    row += 2

    # Paiement
    if invoice.payment_info:
        ws.merge_cells(f'A{row}:{last_col_letter}{row}')
        ws.cell(row=row, column=1, value=invoice.payment_info).font = Font(
            name="Arial", size=9, color=COLORS["DARK_GRAY"], italic=True)
        row += 2

    # ========================================
    # VÉRIFICATION
    # ========================================
    ws.merge_cells(f'A{row}:{last_col_letter}{row}')
    ws.cell(row=row, column=1, value="VÉRIFICATION DES TOTAUX").font = Font(
        name="Arial", bold=True, size=11, color=COLORS["RED"])
    row += 1

    section_str = " + ".join(
        f"{s.expected_total or s.computed_total_before_discounts():,.2f}"
        for s in invoice.sections
    )
    checks = [
        f"Sections: {section_str} = {invoice.sum_section_totals():,.2f} €",
    ]
    if total_disc != 0:
        checks.append(f"Remises: {total_disc:,.2f} €")
    checks.extend([
        f"Total HT: {invoice.sum_section_totals():,.2f} + ({total_disc:,.2f}) = {invoice.total_ht:,.2f} € ✓",
        f"TVA: {' + '.join(f'{t.montant_tva:,.2f}' for t in invoice.tva_rates)} = {invoice.total_tva:,.2f} € ✓",
        f"Total TTC: {invoice.total_ht:,.2f} + {invoice.total_tva:,.2f} = {invoice.total_ttc:,.2f} € ✓",
    ])

    for check in checks:
        ws.merge_cells(f'A{row}:{last_col_letter}{row}')
        ws.cell(row=row, column=1, value=check).font = Font(
            name="Arial", size=8.5, color=COLORS["GREEN_DARK"])
        row += 1

    # Print setup
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    wb.save(output_path)
    print(f"\n📄 Fichier généré: {output_path}")
    return True


# =============================================================================
# POINT D'ENTRÉE PRINCIPAL
# =============================================================================

def process_invoice(invoice, output_path):
    """
    Point d'entrée: vérifie puis génère.
    Retourne True si succès, False si erreur de vérification.
    """
    print("=" * 60)
    print(f"FACTURE: {invoice.numero}")
    print(f"CLIENT: {invoice.client_name}")
    print(f"DATE: {invoice.date}")
    print(f"MAGASIN: {invoice.magasin}")
    print("=" * 60)
    print("\n🔍 Vérification des totaux...")

    success, errors = verify_invoice(invoice)

    if not success:
        print(f"\n❌ GÉNÉRATION ANNULÉE — {len(errors)} erreur(s) détectée(s)")
        print("Corrigez les données et réessayez.")
        return False

    print("\n📊 Génération du fichier Excel...")
    return generate_excel(invoice, output_path)


# =============================================================================
# EXEMPLE D'UTILISATION (template)
# =============================================================================

if __name__ == "__main__":
    print("""
╔══════════════════════════════════════════════════════════╗
║  GÉNÉRATEUR DE FACTURES METRO → EXCEL v3.0              ║
║  Supporte: Metro, Carrefour, Codifrance                 ║
║                                                         ║
║  Usage:                                                 ║
║  1. Créer un objet MetroInvoice()                       ║
║  2. Remplir les sections avec InvoiceItem               ║
║  3. Appeler process_invoice(invoice, "output.xlsx")     ║
║                                                         ║
║  Le système vérifie TOUS les totaux avant génération.   ║
║  Si un montant ne matche pas → pas de fichier généré.   ║
╚══════════════════════════════════════════════════════════╝
    """)
